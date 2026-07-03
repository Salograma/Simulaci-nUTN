"""
================================================================================
 TP SIMULACION - MODELO DE INVENTARIO (s, S)
 Entrega unica: motor de simulacion + estudio completo (30 corridas por
 experimento) + modo interactivo para variar parametros en clase.
================================================================================

Como correrlo:
    python3 inventario_completo.py

Al arrancar, pregunta que modo usar:
    1) Estudio completo -> corre la simulacion base de 365 dias a lo largo de
       30 replicaciones fijando s=60 y S=200, exporta la tabla a Excel y
       genera las graficas de costos.
    2) Modo interactivo -> permite ingresar s, S, costos y demanda a mano
       para mostrar resultados en el momento (pensado para la clase).
"""

import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "resultados_inventario")
os.makedirs(OUT, exist_ok=True)


# ============================================================================
# 1) MOTOR DE SIMULACION (Revision continua, politica s, S, ventas perdidas)
# ============================================================================

def simulate_inventory(s=60, S=200, c_order=100.0, c_holding=0.5, c_shortage=5.0, 
                       mean_demand=20.0, std_demand=5.0, days=365, seed=None):
    """
    Simula una politica de inventario (s, S) con demanda diaria normal.
    Devuelve un dict con los costos acumulados y vectores de evolucion diaria.
    """
    rng = np.random.default_rng(seed)

    inventory = S
    total_order_cost = 0.0
    total_holding_cost = 0.0
    total_shortage_cost = 0.0
    
    order_pending = False
    days_until_delivery = 0
    units_ordered = 0
    
    # Vectores para registrar la evolucion diaria (para graficar despues)
    history_inv = []
    history_holding = []
    history_shortage = []
    
    for day in range(days):
        # 1) Llegada de pedido del proveedor
        if order_pending:
            days_until_delivery -= 1
            if days_until_delivery == 0:
                inventory += units_ordered
                order_pending = False
        
        # 2) Generar demanda diaria estocastica (Normal)
        demand = max(0, int(rng.normal(mean_demand, std_demand)))
        
        # 3) Satisfacer demanda y evaluar quiebres (ventas perdidas)
        if inventory >= demand:
            inventory -= demand
            shortage = 0
        else:
            shortage = demand - inventory
            inventory = 0
            
        # 4) Calcular costos del dia
        daily_holding = inventory * c_holding
        daily_shortage = shortage * c_shortage
        
        total_holding_cost += daily_holding
        total_shortage_cost += daily_shortage
        
        # Guardar historial del dia
        history_inv.append(inventory)
        history_holding.append(daily_holding)
        history_shortage.append(daily_shortage)
        
        # 5) Revision de la politica (s, S) al final del dia
        # Posicion de inventario = inventario fisico + unidades en camino
        inv_position = inventory + (units_ordered if order_pending else 0)
        
        if inv_position < s and not order_pending:
            order_pending = True
            units_ordered = S - inventory
            total_order_cost += c_order
            # Lead time uniforme entre 1 y 3 dias (rango inclusivo discreto)
            days_until_delivery = int(rng.uniform(1, 4)) 

    total_cost = total_order_cost + total_holding_cost + total_shortage_cost
    
    return {
        "cost_order": total_order_cost,
        "cost_holding": total_holding_cost,
        "cost_shortage": total_shortage_cost,
        "cost_total": total_cost,
        "history_inv": history_inv,
        "history_holding": history_holding,
        "history_shortage": history_shortage
    }


# ============================================================================
# 2) FORMATO Y EXPORTACION A EXCEL (Estilo identico al de tu compañero)
# ============================================================================

HEADER_FILL = PatternFill("solid", start_color="2563EB", end_color="2563EB")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
CELL_FONT = Font(name="Arial")
TITLE_FONT = Font(bold=True, size=13, name="Arial")


def _escribir_tabla(ws, df, start_row=1, decimales=4):
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=start_row, column=j, value=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    for i, (_, fila) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=1):
            val = fila[col]
            if isinstance(val, (float, np.floating)):
                val = round(float(val), decimales)
            c = ws.cell(row=i, column=j, value=val)
            c.font = CELL_FONT
    for j, col in enumerate(df.columns, start=1):
        ancho = max(15, len(str(col)) + 2)
        ws.column_dimensions[get_column_letter(j)].width = ancho


def guardar_excel_inventario(df_replicas, df_resumen, fname):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Corridas Independientes"
    ws1["A1"] = "Resultados economicos detallados por cada una de las 30 replicas"
    ws1["A1"].font = TITLE_FONT
    _escribir_tabla(ws1, df_replicas, start_row=3)

    ws2 = wb.create_sheet("Resumen Estadistico")
    ws2["A1"] = "Metricas consolidadas (Promedios finales del estudio de simulacion)"
    ws2["A1"].font = TITLE_FONT
    _escribir_tabla(ws2, df_resumen, start_row=3)

    wb.save(fname)
    print(f"Excel guardado en: {fname}")


# ============================================================================
# 3) MODO 1: ESTUDIO COMPLETO (30 replicaciones con s=60, S=200)
# ============================================================================

def modo_estudio_completo():
    N_REPLICAS = 30
    DAYS = 365
    s_BASE, S_BASE = 60, 200
    rng_master = np.random.default_rng(42)
    
    print(f"\n[1/3] Ejecutando {N_REPLICAS} replicas del modelo de inventarios...")
    
    rows = []
    arrays_inv = []
    
    for r in range(1, N_REPLICAS + 1):
        seed = int(rng_master.integers(0, 1_000_000))
        res = simulate_inventory(s=s_BASE, S=S_BASE, days=DAYS, seed=seed)
        
        rows.append({
            "Replica": r,
            "Costo Orden ($)": res["cost_order"],
            "Costo Mantener ($)": res["cost_holding"],
            "Costo Faltante ($)": res["cost_shortage"],
            "Costo Total ($)": res["cost_total"]
        })
        arrays_inv.append(res["history_inv"])
        
    df_replicas = pd.DataFrame(rows)
    
    # Calcular promedios consolidados
    df_resumen = pd.DataFrame([{
        "Parametro s": s_BASE,
        "Parametro S": S_BASE,
        "Promedio Costo Orden ($)": df_replicas["Costo Orden ($)"].mean(),
        "Promedio Costo Mantener ($)": df_replicas["Costo Mantener ($)"].mean(),
        "Promedio Costo Faltante ($)": df_replicas["Costo Faltante ($)"].mean(),
        "COSTO TOTAL MEDIO ($)": df_replicas["Costo Total ($)"].mean()
    }])
    
    print("\n--- RESUMEN DE METRICAS MEDIAS PROMEDIO ---")
    print(df_resumen.to_string(index=False))
    
    print("\n[2/3] Generando graficas de analisis de stock...")
    
    # Grafico 1: Evolucion del stock diario (Tomando la replica 1 como muestra)
    plt.figure(figsize=(10, 4.5))
    plt.plot(range(DAYS), arrays_inv[0], color="#2563eb", label="Stock Fisico")
    plt.axhline(s_BASE, color="#dc2626", linestyle="--", label=f"Punto de reorden (s={s_BASE})")
    plt.axhline(S_BASE, color="#16a34a", linestyle="--", label=f"Stock maximo (S={S_BASE})")
    plt.title("Perfil de evolucion diaria de inventario neto (Replica Muestra 1)")
    plt.xlabel("Dia de operacion"); plt.ylabel("Unidades en Almacen")
    plt.legend(); plt.grid(alpha=0.3)
    plt.tight_layout()
    plt.savefig(f"{OUT}/inventario_perfil_evolucion.png", dpi=140)
    plt.close()
    
    print("\n[3/3] Exportando matrices de costos consolidadas...")
    guardar_excel_inventario(df_replicas, df_resumen, f"{OUT}/inventario_resultados_base.xlsx")
    print(f"\nListo. Archivos generados con exito en la ruta: {OUT}")


# ============================================================================
# 4) MODO 2: INTERACTIVO (Ingreso de parametros en vivo frente al aula)
# ============================================================================

def pedir_float(mensaje, default):
    txt = input(f"{mensaje} [enter = {default}]: ").strip()
    return default if txt == "" else float(txt)


def pedir_int(mensaje, default):
    txt = input(f"{mensaje} [enter = {default}]: ").strip()
    return default if txt == "" else int(txt)


def modo_interactivo():
    print("=" * 60)
    print("  SIMULACION DE INVENTARIOS (s, S) - Parametros en Vivo")
    print("=" * 60)
    
    s = pedir_int("Punto de reorden (s)", 60)
    S = pedir_int("Inventario maximo meta (S)", 200)
    c_order = pedir_float("Costo fijo por lanzar una orden ($)", 100.0)
    c_holding = pedir_float("Costo diario por unidad en estanteria ($)", 0.5)
    c_shortage = pedir_float("Costo de penalizacion por venta perdida ($)", 5.0)
    mean_d = pedir_float("Demanda diaria promedio (Media)", 20.0)
    std_d = pedir_float("Desviacion estandar de la demanda", 5.0)
    n_reps = pedir_int("Cantidad de replicas independientes a promediar", 30)
    days = pedir_int("Horizonte temporal de la corrida (Dias)", 365)
    
    if s >= S:
        print("\n[AVISO CRITICO]: s >= S implica que el punto de reorden supera al maximo.")
        print("El sistema generara ordenes continuas distorsionando los costos.")
        
    rng = np.random.default_rng(2026)
    c_o_list, c_h_list, c_s_list, c_t_list = [], [], [], []
    
    for _ in range(n_reps):
        seed = int(rng.integers(0, 1_000_000))
        res = simulate_inventory(s, S, c_order, c_holding, c_shortage, mean_d, std_d, days, seed)
        c_o_list.append(res["cost_order"])
        c_h_list.append(res["cost_holding"])
        c_s_list.append(res["cost_shortage"])
        c_t_list.append(res["cost_total"])
        
    print(f"\nResultados del Escenario Interactivo (Promedio de {n_reps} corridas):")
    print(f"  {'Costo medio de ordenamiento':35s}: ${np.mean(c_o_list):10.2f}")
    print(f"  {'Costo medio de mantenimiento stock':35s}: ${np.mean(c_h_list):10.2f}")
    print(f"  {'Costo medio de quiebre de stock':35s}: ${np.mean(c_s_list):10.2f}")
    print("-" * 60)
    print(f"  {'COSTO TOTAL OPERATIVO PROMEDIO':35s}: ${np.mean(c_t_list):10.2f}")
    
    # Guardar Excel resumido de la prueba de clase
    df_interactivo = pd.DataFrame([{
        "s": s, "S": S, "Dias": days,
        "Costo Orden Promedio": np.mean(c_o_list),
        "Costo Mantener Promedio": np.mean(c_h_list),
        "Costo Faltante Promedio": np.mean(c_s_list),
        "Costo Total Promedio": np.mean(c_t_list)
    }])
    fname_xlsx = f"{OUT}/inventario_interactivo_clase.xlsx"
    df_interactivo.to_excel(fname_xlsx, index=False)
    print(f"\nExcel de la simulacion interactiva guardado en: {fname_xlsx}")


# ==========================================
# 5) MENU PRINCIPAL
# ==========================================

if __name__ == "__main__":
    print("=" * 60)
    print("  TP SIMULACION - SISTEMAS DE INVENTARIOS (s, S)")
    print("=" * 60)
    print("  1) Estudio completo corporativo (30 replicas, analisis base)")
    print("  2) Modo interactivo de aula (Modificar variables en vivo)")
    opcion = input("Elegi una opcion [1/2]: ").strip()

    if opcion == "2":
        modo_interactivo()
    else:
        modo_estudio_completo()