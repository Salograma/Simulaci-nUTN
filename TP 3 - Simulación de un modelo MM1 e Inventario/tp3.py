"""
================================================================================
 TP SIMULACION - MODELO M/M/1 y M/M/1/K
 Entrega unica: motor de simulacion + estudio completo (30 corridas por
 experimento) + modo interactivo para variar parametros en clase.
================================================================================

Como correrlo:
    python3 mm1_completo.py

Al arrancar, pregunta que modo usar:
    1) Estudio completo -> corre TODOS los experimentos pedidos por la
       catedra (variacion de rho 25/50/75/100/125%, cola finita
       K=0,2,5,10,50, 30 corridas cada uno) y guarda tablas (Excel) +
       graficos.
    2) Modo interactivo -> permite ingresar lambda, mu, K, etc. a mano
       para mostrar resultados en el momento (pensado para la clase).
"""

import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "resultados_mm1")
os.makedirs(OUT, exist_ok=True)


# ============================================================================
# 1) MOTOR DE SIMULACION (eventos discretos, un servidor, FIFO)
# ============================================================================

def simulate_mm1k(lam, mu, K=None, sim_time=20000, warmup=2000, seed=None):
    """
    Simula una cola M/M/1/K (K = capacidad de la COLA; capacidad total del
    sistema = K+1, contando al cliente en servicio). K=None => cola infinita.
    Devuelve un dict con las metricas de la corrida (descartando el warmup).
    """
    rng = np.random.default_rng(seed)

    clock = 0.0
    n_system = 0
    next_arrival = rng.exponential(1.0 / lam)
    next_departure = np.inf

    area_L = 0.0
    area_Lq = 0.0
    busy_time = 0.0
    last_event_time = 0.0
    warmup_done = False

    n_arrivals = 0
    n_blocked = 0

    arrival_times_in_system = []
    sojourn_times = []

    max_track = (K + 2) if K is not None else 60
    time_in_state = np.zeros(max_track + 5)

    def record_state_time(dt, n_before):
        nonlocal area_L, area_Lq, busy_time
        if warmup_done and dt > 0:
            area_L += n_before * dt
            nq = max(n_before - 1, 0)
            area_Lq += nq * dt
            if n_before > 0:
                busy_time += dt
            idx = min(n_before, max_track + 4)
            time_in_state[idx] += dt

    while clock < sim_time:
        if not warmup_done and clock >= warmup:
            warmup_done = True
            last_event_time = clock

        t_next = min(next_arrival, next_departure)
        dt = t_next - last_event_time
        record_state_time(dt, n_system)

        clock = t_next
        last_event_time = clock

        if next_arrival <= next_departure:
            n_arrivals += 1
            capacity = np.inf if K is None else (K + 1)
            if n_system < capacity:
                n_system += 1
                if n_system == 1:
                    next_departure = clock + rng.exponential(1.0 / mu)
                if warmup_done:
                    arrival_times_in_system.append(clock)
            else:
                n_blocked += 1
            next_arrival = clock + rng.exponential(1.0 / lam)
        else:
            n_system -= 1
            if warmup_done and arrival_times_in_system:
                t_in = arrival_times_in_system.pop(0)
                sojourn_times.append(clock - t_in)
            if n_system > 0:
                next_departure = clock + rng.exponential(1.0 / mu)
            else:
                next_departure = np.inf

    post_warmup_time = clock - warmup if clock > warmup else 1e-9

    L = area_L / post_warmup_time
    Lq = area_Lq / post_warmup_time
    util = busy_time / post_warmup_time
    W = np.mean(sojourn_times) if sojourn_times else np.nan
    Wq = max(W - 1.0 / mu, 0) if not np.isnan(W) else np.nan
    p_block = n_blocked / n_arrivals if n_arrivals > 0 else 0.0
    dist = time_in_state / post_warmup_time

    return {"L": L, "Lq": Lq, "W": W, "Wq": Wq, "util": util,
            "p_block": p_block, "n_arrivals": n_arrivals,
            "n_blocked": n_blocked, "dist_n": dist}


# ============================================================================
# 2) FORMULAS TEORICAS (para comparar contra la simulacion)
# ============================================================================

def theoretical_mm1(lam, mu):
    """Formulas clasicas de M/M/1 con cola INFINITA. Solo validas si rho<1."""
    rho = lam / mu
    if rho >= 1:
        return {"rho": rho, "L": np.inf, "Lq": np.inf, "W": np.inf, "Wq": np.inf}
    L = rho / (1 - rho)
    Lq = rho ** 2 / (1 - rho)
    W = L / lam
    Wq = Lq / lam
    return {"rho": rho, "L": L, "Lq": Lq, "W": W, "Wq": Wq}


def theoretical_pn(lam, mu, n_max=15):
    """Distribucion teorica P(n) para cola infinita: Pn = (1-rho)*rho^n."""
    rho = lam / mu
    return [(1 - rho) * rho ** n for n in range(n_max + 1)]


def theoretical_mm1k_full(lam, mu, K):
    """
    Formulas teoricas COMPLETAS de M/M/1/K (cola finita).
    K = capacidad de la cola; N = K+1 = capacidad total del sistema.
    Valida para cualquier rho (incluso rho>=1), porque con capacidad finita
    el sistema siempre es estable.
    """
    rho = lam / mu
    N = K + 1
    n_vals = np.arange(0, N + 1)
    if abs(rho - 1.0) < 1e-9:
        Pn = np.full(N + 1, 1.0 / (N + 1))
    else:
        p0 = (1 - rho) / (1 - rho ** (N + 1))
        Pn = p0 * rho ** n_vals

    p_block = Pn[-1]
    lam_eff = lam * (1 - p_block)
    util = 1 - Pn[0]
    L = float(np.sum(n_vals * Pn))
    Lq = L - util
    W = L / lam_eff if lam_eff > 0 else np.nan
    Wq = Lq / lam_eff if lam_eff > 0 else np.nan

    return {"rho": rho, "L": L, "Lq": Lq, "W": W, "Wq": Wq,
            "util": util, "p_block": p_block, "Pn": Pn}


# ============================================================================
# 3) EXPORTAR RESULTADOS A EXCEL (formato prolijo, no CSV plano)
# ============================================================================

HEADER_FILL = PatternFill("solid", start_color="2563EB", end_color="2563EB")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
CELL_FONT = Font(name="Arial")
TITLE_FONT = Font(bold=True, size=13, name="Arial")


def _escribir_tabla(ws, df, start_row=1, decimales=4):
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=start_row, column=j, value=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    for i, (_, fila) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=1):
            val = fila[col]
            if isinstance(val, (float, np.floating)):
                val = round(float(val), decimales)
            c = ws.cell(row=i, column=j, value=val)
            c.font = CELL_FONT
    for j, col in enumerate(df.columns, start=1):
        ancho = max(12, len(str(col)) + 2)
        ws.column_dimensions[get_column_letter(j)].width = ancho


def guardar_excel_mm1(df_metricas, df_bloqueo, fname):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Metricas M-M-1"
    ws1["A1"] = "Medidas de rendimiento M/M/1 - variacion de rho (25% a 125% de mu)"
    ws1["A1"].font = TITLE_FONT
    _escribir_tabla(ws1, df_metricas, start_row=3)

    ws2 = wb.create_sheet("Bloqueo M-M-1-K")
    ws2["A1"] = "Probabilidad de denegacion de servicio - cola finita (K=0,2,5,10,50)"
    ws2["A1"].font = TITLE_FONT
    _escribir_tabla(ws2, df_bloqueo, start_row=3)

    wb.save(fname)
    print(f"Excel guardado en: {fname}")


# ============================================================================
# 4) MODO 1: ESTUDIO COMPLETO (lo que pide la catedra)
#    - Variacion de rho: 25/50/75/100/125% de mu
#    - 30 corridas por experimento (se reporta el PROMEDIO de las 30)
#    - Cola finita K=0,2,5,10,50 -> probabilidad de bloqueo
#    - Graficos: comparacion teorico/simulado, convergencia temporal,
#      distribucion P(n), bloqueo vs K
# ============================================================================

def modo_estudio_completo():
    MU = 1.0
    RHO_LEVELS = [0.25, 0.50, 0.75, 1.00, 1.25]
    N_REPLICAS = 30
    SIM_TIME = 20000
    WARMUP = 2000
    K_LEVELS = [0, 2, 5, 10, 50]
    rng_master = np.random.default_rng(42)

    print("\n[1/4] Corriendo experimento de cola infinita (variando rho)...")
    rows = []
    for rho in RHO_LEVELS:
        lam = rho * MU
        L_list, Lq_list, W_list, Wq_list, util_list = [], [], [], [], []
        for _ in range(N_REPLICAS):
            seed = int(rng_master.integers(0, 1_000_000))
            res = simulate_mm1k(lam, MU, K=None if rho < 1 else 5000,
                                 sim_time=SIM_TIME, warmup=WARMUP, seed=seed)
            L_list.append(res["L"]); Lq_list.append(res["Lq"])
            W_list.append(res["W"]); Wq_list.append(res["Wq"])
            util_list.append(res["util"])

        teo = theoretical_mm1(lam, MU)
        rows.append({
            "rho_%": int(rho * 100), "lambda": lam, "mu": MU,
            "L_sim": np.mean(L_list), "L_teo": teo["L"],
            "Lq_sim": np.mean(Lq_list), "Lq_teo": teo["Lq"],
            "W_sim": np.mean(W_list), "W_teo": teo["W"],
            "Wq_sim": np.mean(Wq_list), "Wq_teo": teo["Wq"],
            "util_sim": np.mean(util_list), "util_teo": min(rho, 1.0),
        })

    df1 = pd.DataFrame(rows)
    print(df1.to_string(index=False))

    fig, axes = plt.subplots(2, 2, figsize=(11, 8))
    metrics = [("L", "Clientes promedio en sistema (L)"),
               ("Lq", "Clientes promedio en cola (Lq)"),
               ("W", "Tiempo promedio en sistema (W)"),
               ("Wq", "Tiempo promedio en cola (Wq)")]
    for ax, (key, title) in zip(axes.flat, metrics):
        x = df1["rho_%"]
        ax.plot(x, df1[f"{key}_sim"], "o-", label="Simulado", color="#2563eb")
        y_teo = df1[f"{key}_teo"].replace([np.inf], np.nan)
        ax.plot(x, y_teo, "s--", label="Teorico M/M/1", color="#dc2626")
        ax.set_title(title); ax.set_xlabel("rho (%)")
        ax.legend(fontsize=8); ax.grid(alpha=0.3)
    plt.tight_layout()
    plt.savefig(f"{OUT}/mm1_comparacion_teorico_vs_simulado.png", dpi=140)
    plt.close()

    plt.figure(figsize=(6, 4.5))
    plt.plot(df1["rho_%"], df1["util_sim"], "o-", label="Simulado", color="#2563eb")
    plt.plot(df1["rho_%"], df1["util_teo"], "s--", label="Teorico", color="#dc2626")
    plt.title("Utilizacion del servidor")
    plt.xlabel("rho (%)"); plt.ylabel("Utilizacion")
    plt.legend(); plt.grid(alpha=0.3)
    plt.tight_layout()
    plt.savefig(f"{OUT}/mm1_utilizacion.png", dpi=140)
    plt.close()

    print("\n[2/4] Generando grafico de convergencia temporal...")

    def simulate_trace(lam, mu, sim_time, seed):
        rng = np.random.default_rng(seed)
        clock = 0.0; n = 0
        next_arrival = rng.exponential(1 / lam); next_departure = np.inf
        times = [0.0]; Ls = [0]
        while clock < sim_time:
            t_next = min(next_arrival, next_departure)
            clock = t_next
            if next_arrival <= next_departure:
                n += 1
                if n == 1:
                    next_departure = clock + rng.exponential(1 / mu)
                next_arrival = clock + rng.exponential(1 / lam)
            else:
                n -= 1
                next_departure = clock + rng.exponential(1 / mu) if n > 0 else np.inf
            times.append(clock); Ls.append(n)
        return np.array(times), np.array(Ls)

    def running_avg(times, Ls):
        cum_area = np.concatenate([[0], np.cumsum(np.diff(times) * Ls[:-1])])
        with np.errstate(divide="ignore", invalid="ignore"):
            return np.where(times > 0, cum_area / times, 0)

    plt.figure(figsize=(9, 5))
    for rho in [0.5, 0.75, 0.9]:
        t, Ls = simulate_trace(rho * MU, MU, sim_time=3000, seed=1)
        avg = running_avg(t, Ls)
        plt.plot(t, avg, label=f"rho={int(rho*100)}%")
    plt.title("Convergencia de L promedio acumulado vs tiempo de simulacion")
    plt.xlabel("Tiempo de simulacion"); plt.ylabel("L promedio acumulado")
    plt.legend(); plt.grid(alpha=0.3)
    plt.tight_layout()
    plt.savefig(f"{OUT}/mm1_convergencia_temporal.png", dpi=140)
    plt.close()

    print("\n[3/4] Generando distribucion P(n) para rho=75%...")
    rho_demo = 0.75
    res_demo = simulate_mm1k(rho_demo * MU, MU, K=None, sim_time=SIM_TIME,
                              warmup=WARMUP, seed=7)
    dist_sim = res_demo["dist_n"][:12]
    dist_teo = theoretical_pn(rho_demo * MU, MU, n_max=11)

    plt.figure(figsize=(7, 4.5))
    xw = np.arange(12)
    plt.bar(xw - 0.2, dist_sim, width=0.4, label="Simulado", color="#2563eb")
    plt.bar(xw + 0.2, dist_teo, width=0.4, label="Teorico", color="#dc2626")
    plt.title(f"Distribucion P(n clientes en sistema) - rho={int(rho_demo*100)}%")
    plt.xlabel("n"); plt.ylabel("Probabilidad")
    plt.legend(); plt.grid(alpha=0.3)
    plt.tight_layout()
    plt.savefig(f"{OUT}/mm1_distribucion_n.png", dpi=140)
    plt.close()

    print("\n[4/4] Corriendo experimentos de cola finita (bloqueo)...")
    rows_block = []
    for rho in RHO_LEVELS:
        lam = rho * MU
        for K in K_LEVELS:
            pb_list = []
            for _ in range(N_REPLICAS):
                seed = int(rng_master.integers(0, 1_000_000))
                res = simulate_mm1k(lam, MU, K=K, sim_time=SIM_TIME,
                                     warmup=WARMUP, seed=seed)
                pb_list.append(res["p_block"])
            teo = theoretical_mm1k_full(lam, MU, K)
            rows_block.append({
                "rho_%": int(rho * 100), "K": K,
                "p_block_sim": np.mean(pb_list),
                "p_block_teo": teo["p_block"]
            })

    df2 = pd.DataFrame(rows_block)
    print(df2.to_string(index=False))

    plt.figure(figsize=(8, 5.5))
    for rho in RHO_LEVELS:
        sub = df2[df2["rho_%"] == int(rho * 100)]
        plt.plot(sub["K"], sub["p_block_sim"], "o-", label=f"rho={int(rho*100)}% (sim)")
    plt.title("Probabilidad de denegacion de servicio vs capacidad de cola (K)")
    plt.xlabel("K (capacidad de la cola)"); plt.ylabel("P(bloqueo)")
    plt.legend(fontsize=8); plt.grid(alpha=0.3)
    plt.tight_layout()
    plt.savefig(f"{OUT}/mm1_bloqueo_vs_K.png", dpi=140)
    plt.close()

    guardar_excel_mm1(df1, df2, f"{OUT}/mm1_resultados.xlsx")
    print(f"\nListo. Archivos generados en: {OUT}")


# ============================================================================
# 5) MODO 2: INTERACTIVO (para mostrar en clase, variando parametros)
# ============================================================================

def pedir_float(mensaje, default):
    txt = input(f"{mensaje} [enter = {default}]: ").strip()
    return default if txt == "" else float(txt)


def pedir_int(mensaje, default):
    txt = input(f"{mensaje} [enter = {default}]: ").strip()
    return default if txt == "" else int(txt)


def pedir_K(mensaje):
    txt = input(f"{mensaje} [enter = infinita]: ").strip()
    return None if txt == "" else int(txt)


def modo_interactivo():
    print("=" * 60)
    print("  SIMULACION M/M/1 - Ingreso de parametros")
    print("=" * 60)

    lam = pedir_float("Tasa de arribo (lambda, clientes/hora)", 7.5)
    mu = pedir_float("Tasa de servicio (mu, clientes/hora)", 10.0)
    K = pedir_K("Capacidad de la cola (K, lugares para esperar)")
    n_reps = pedir_int("Cantidad de corridas (minimo 30 sugerido)", 30)
    sim_time = pedir_float("Tiempo de simulacion por corrida (horas)", 2000)
    warmup = pedir_float("Tiempo de warmup a descartar (horas)", 200)

    rho = lam / mu
    print(f"\n--> rho = lambda/mu = {rho:.3f}")
    if rho >= 1 and K is None:
        print("    OJO: rho >= 1 con cola infinita -> sistema inestable.")

    rng = np.random.default_rng(2024)
    L_l, Lq_l, W_l, Wq_l, u_l, pb_l = [], [], [], [], [], []
    for _ in range(n_reps):
        seed = int(rng.integers(0, 1_000_000))
        res = simulate_mm1k(lam, mu, K=K, sim_time=sim_time, warmup=warmup, seed=seed)
        L_l.append(res["L"]); Lq_l.append(res["Lq"])
        W_l.append(res["W"]); Wq_l.append(res["Wq"])
        u_l.append(res["util"]); pb_l.append(res["p_block"])

    def resumen(nombre, x):
        m = float(np.nanmean(np.array(x, dtype=float)))
        print(f"  {nombre:35s}: {m:10.4f}")
        return m

    print(f"\nResultados de la simulacion (promedio de {n_reps} corridas):")
    L_m = resumen("L  (clientes en el sistema)", L_l)
    Lq_m = resumen("Lq (clientes en cola)", Lq_l)
    W_m = resumen("W  (tiempo en el sistema)", W_l)
    Wq_m = resumen("Wq (tiempo en cola)", Wq_l)
    u_m = resumen("Utilizacion del servidor", u_l)
    pb_m = resumen("P(bloqueo)", pb_l) if K is not None else None

    print("\nComparacion contra valor teorico:")
    fila_resultado = None
    if K is not None:
        teo = theoretical_mm1k_full(lam, mu, K)
        print("  (usando formula M/M/1/K, porque la cola es finita)")
        print(f"  {'Metrica':12s}{'Simulado':>14s}{'Teorico':>14s}{'Dif %':>10s}")
        metricas = [("L", L_m, teo["L"]), ("Lq", Lq_m, teo["Lq"]),
                    ("W", W_m, teo["W"]), ("Wq", Wq_m, teo["Wq"]),
                    ("Utilizacion", u_m, teo["util"]),
                    ("P(bloqueo)", pb_m, teo["p_block"])]
        for nombre, sim_v, teo_v in metricas:
            diff = 100 * (sim_v - teo_v) / teo_v if abs(teo_v) > 1e-9 else float("nan")
            diff_txt = f"{diff:8.2f}%" if not np.isnan(diff) else "  (n/a)"
            print(f"  {nombre:12s}{sim_v:14.4g}{teo_v:14.4g}{diff_txt:>10s}")
        fila_resultado = {m[0]: {"Simulado": m[1], "Teorico": m[2]} for m in metricas}
    elif rho < 1:
        teo = theoretical_mm1(lam, mu)
        print("  (usando formula M/M/1 de cola infinita)")
        print(f"  {'Metrica':12s}{'Simulado':>14s}{'Teorico':>14s}{'Dif %':>10s}")
        metricas = [("L", L_m, teo["L"]), ("Lq", Lq_m, teo["Lq"]),
                    ("W", W_m, teo["W"]), ("Wq", Wq_m, teo["Wq"]),
                    ("Utilizacion", u_m, rho)]
        for nombre, sim_v, teo_v in metricas:
            diff = 100 * (sim_v - teo_v) / teo_v if teo_v != 0 else float("nan")
            print(f"  {nombre:12s}{sim_v:14.4f}{teo_v:14.4f}{diff:9.2f}%")
        fila_resultado = {m[0]: {"Simulado": m[1], "Teorico": m[2]} for m in metricas}
    else:
        print("  rho >= 1 y cola infinita: no existe formula teorica en estado")
        print("  estable (el sistema es inestable, la cola crece sin limite).")

    if fila_resultado is not None:
        df_res = pd.DataFrame(fila_resultado).T.reset_index().rename(columns={"index": "Metrica"})
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultado interactivo"
        ws["A1"] = f"M/M/1 interactivo - lambda={lam}, mu={mu}, K={'infinita' if K is None else K}"
        ws["A1"].font = TITLE_FONT
        _escribir_tabla(ws, df_res, start_row=3)
        fname_xlsx = f"{OUT}/mm1_interactivo_resultado.xlsx"
        wb.save(fname_xlsx)
        print(f"\nExcel guardado en: {fname_xlsx}")

    res_demo = simulate_mm1k(lam, mu, K=K, sim_time=sim_time, warmup=warmup, seed=99)
    if K is not None:
        n_max = K + 1
        dist_sim = res_demo["dist_n"][:n_max + 1]
        dist_teo = theoretical_mm1k_full(lam, mu, K)["Pn"]
    else:
        n_max = min(15, len(res_demo["dist_n"]) - 1)
        dist_sim = res_demo["dist_n"][:n_max + 1]
        dist_teo = theoretical_pn(lam, mu, n_max=n_max) if rho < 1 else [np.nan] * (n_max + 1)

    plt.figure(figsize=(7, 4.5))
    xw = np.arange(n_max + 1)
    plt.bar(xw - 0.2, dist_sim, width=0.4, label="Simulado", color="#2563eb")
    plt.bar(xw + 0.2, dist_teo, width=0.4, label="Teorico", color="#dc2626")
    plt.title(f"P(n clientes en sistema) - lambda={lam}, mu={mu}, rho={rho:.2f}")
    plt.xlabel("n"); plt.ylabel("Probabilidad")
    plt.legend(); plt.grid(alpha=0.3)
    plt.tight_layout()
    fname = f"{OUT}/mm1_interactivo_distribucion_n.png"
    plt.savefig(fname, dpi=140)
    plt.close()
    print(f"Grafico guardado en: {fname}")

    rng2 = np.random.default_rng(int(rng.integers(0, 1_000_000)))
    clock = 0.0; n = 0
    next_a = rng2.exponential(1 / lam); next_d = np.inf
    times = [0.0]; Ls = [0]
    cap = np.inf if K is None else K + 1
    t_max = min(sim_time, 5000)
    while clock < t_max:
        t_next = min(next_a, next_d)
        clock = t_next
        if next_a <= next_d:
            if n < cap:
                n += 1
                if n == 1:
                    next_d = clock + rng2.exponential(1 / mu)
            next_a = clock + rng2.exponential(1 / lam)
        else:
            n -= 1
            next_d = clock + rng2.exponential(1 / mu) if n > 0 else np.inf
        times.append(clock); Ls.append(n)
    times = np.array(times); Ls = np.array(Ls)
    cum_area = np.concatenate([[0], np.cumsum(np.diff(times) * Ls[:-1])])
    with np.errstate(divide="ignore", invalid="ignore"):
        avg = np.where(times > 0, cum_area / times, 0)

    plt.figure(figsize=(8, 4.5))
    plt.plot(times, avg, color="#2563eb")
    if rho < 1 and K is None:
        plt.axhline(theoretical_mm1(lam, mu)["L"], color="#dc2626", ls="--", label="L teorico")
        plt.legend()
    plt.title("Convergencia de L promedio acumulado vs tiempo de simulacion")
    plt.xlabel("Tiempo"); plt.ylabel("L promedio acumulado")
    plt.grid(alpha=0.3)
    plt.tight_layout()
    fname2 = f"{OUT}/mm1_interactivo_convergencia.png"
    plt.savefig(fname2, dpi=140)
    plt.close()
    print(f"Grafico guardado en: {fname2}")


# ============================================================================
# 6) MENU PRINCIPAL
# ============================================================================

if __name__ == "__main__":
    print("=" * 60)
    print("  TP SIMULACION - MODELO M/M/1")
    print("=" * 60)
    print("  1) Estudio completo (todos los experimentos, 30 corridas c/u)")
    print("  2) Modo interactivo (ingresar parametros a mano, para clase)")
    opcion = input("Elegi una opcion [1/2]: ").strip()

    if opcion == "2":
        modo_interactivo()
    else:
        modo_estudio_completo()